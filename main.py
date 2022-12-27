from telegram.ext import *
import requests
import openpyxl
import keys
import datetime

print('Bot is UP!')


# Lets us use the /start command
# def start_command(update, context):
#     update.message.reply_text('Hello there! I\'m a bot. What\'s up?')

# Lets see my btc holdings with /mybtc command
def my_btc(update, context):
    # Use the CoinMarketCap API to get the current price of Bitcoin
    api_key = keys.COINMARKETCAP_API_KEY
    response = requests.get(
        'https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest',
        headers={'X-CMC_PRO_API_KEY': api_key},
        params={'start': '1', 'limit': '1', 'convert': 'EUR'}
    )
    data = response.json()

    # Get the current price of Bitcoin in USD
    bitcoin_price = data['data'][0]['quote']['EUR']['price']

    # Send the price back to the user via the bot
    #update.message.reply_text(f'The current price of Bitcoin is €{bitcoin_price:.2f}')

    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook('holdings.xlsx')
    sheet = wb['Sheet1']

    # Find the cell containing the BTC holdings
    for row in sheet.rows:
        for cell in row:
            if cell.value == 'BTC':
                btc_holdings = cell.offset(0, 1).value
                break
    
    final_btc_holdings = btc_holdings * bitcoin_price

    update.message.reply_text(f'Bitcoin Price is €{bitcoin_price:.2f} and your BTC is worth €{final_btc_holdings:.2f}')

# Lets see my btc holdings with /mybtc command
def my_rvn(update, context):
    # Use the CoinMarketCap API to get the current price of RVN
    api_key = keys.COINMARKETCAP_API_KEY
    response = requests.get(
        'https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest',
        headers={'X-CMC_PRO_API_KEY': api_key},
        params={'start': '102', 'limit': '102', 'convert': 'USD'}
    )
    data = response.json()

    # Get the current price of Bitcoin in USD
    ravencoin_price = data['data'][0]['quote']['USD']['price']

    # Send the price back to the user via the bot
    # update.message.reply_text(f'The current price of Ravencoin is €{ravencoin_price:.2f}')

    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook('holdings.xlsx')
    sheet = wb['Sheet1']

    # Find the cell containing the BTC holdings
    for row in sheet.rows:
        for cell in row:
            if cell.value == 'RVN':
                rvn_holdings = cell.offset(0, 1).value
                break
    
    final_rvn_holdings = rvn_holdings * ravencoin_price

    update.message.reply_text(f'Ravencoin Price is €{ravencoin_price:.2f} and your BTC is worth €{final_rvn_holdings:.2f}')


#daily notification

def callback_auto_message(context):
    # Use the CoinMarketCap API to get the current price of RVN
    api_key = keys.COINMARKETCAP_API_KEY
    response = requests.get(
        'https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest',
        headers={'X-CMC_PRO_API_KEY': api_key},
        params={'start': '102', 'limit': '102', 'convert': 'USD'}
    )
    data = response.json()

    # Get the current price of Bitcoin in USD
    ravencoin_price = data['data'][0]['quote']['USD']['price']

    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook('holdings.xlsx')
    sheet = wb['Sheet1']

    # Find the cell containing the BTC holdings
    for row in sheet.rows:
        for cell in row:
            if cell.value == 'RVN':
                rvn_holdings = cell.offset(0, 1).value
                break
    
    final_rvn_holdings = rvn_holdings * ravencoin_price

     # Use the CoinMarketCap API to get the current price of Bitcoin
    api_key = keys.COINMARKETCAP_API_KEY
    response = requests.get(
        'https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest',
        headers={'X-CMC_PRO_API_KEY': api_key},
        params={'start': '1', 'limit': '1', 'convert': 'EUR'}
    )
    data = response.json()

    # Get the current price of Bitcoin in USD
    bitcoin_price = data['data'][0]['quote']['EUR']['price']

    # Send the price back to the user via the bot
    #update.message.reply_text(f'The current price of Bitcoin is €{bitcoin_price:.2f}')

    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook('holdings.xlsx')
    sheet = wb['Sheet1']

    # Find the cell containing the BTC holdings
    for row in sheet.rows:
        for cell in row:
            if cell.value == 'BTC':
                btc_holdings = cell.offset(0, 1).value
                break
    
    final_btc_holdings = btc_holdings * bitcoin_price

    context.bot.send_message(chat_id='1987606009', text=f'Your Ravencoin holdings are now worth €{final_rvn_holdings:.2f} while the price of Ravencoin is {ravencoin_price:.2f} and your are holding {rvn_holdings} RVN')
    context.bot.send_message(chat_id='1987606009', text=f'Your Bitcoin holdings are now worth €{final_btc_holdings:.2f} while the price of Bitcoin is {bitcoin_price:.2f} and your are holding {btc_holdings} BTC')


def start_auto_messaging(update, context):
    chat_id = update.message.chat_id
    context.job_queue.run_repeating(callback_auto_message, 10, context=chat_id, name=str(chat_id))
    # context.job_queue.run_once(callback_auto_message, 1, context=chat_id)
    # context.job_queue.run_daily(callback_auto_message, time=datetime.time(hour=21, minute=50), days=(0, 1, 2, 3, 4, 5, 6), context=chat_id)

def stop_notify(update, context):
    chat_id = update.message.chat_id
    context.bot.send_message(chat_id=chat_id, text='Stopping automatic messages!')
    job = context.job_queue.get_jobs_by_name(str(chat_id))
    job[0].schedule_removal()





# Lets us use the /help command
# def help_command(update, context):
#     update.message.reply_text('Try typing anything and I will do my best to respond!')


# Lets us use the /custom command
# def custom_command(update, context):
#     update.message.reply_text('This is a custom command, you can add whatever text you want here.')


def handle_response(text) -> str:
    # Create your own response logic

    if 'hello' in text:
        return 'Hey there!'

    if 'how are you' in text:
        return 'I\'m good!'

    return 'I don\'t understand'


def handle_message(update, context):
    # Get basic info of the incoming message
    message_type = update.message.chat.type
    text = str(update.message.text).lower()
    response = ''

    # Print a log for debugging
    print(f'User ({update.message.chat.id}) says: "{text}" in: {message_type}')

    # React to group messages only if users mention the bot directly
    if message_type == 'group':
        # Replace with your bot username
        if '@bot19292bot' in text:
            new_text = text.replace('@bot19292bot', '').strip()
            response = handle_response(new_text)
    else:
        response = handle_response(text)

    # Reply normal if the message is in private
    update.message.reply_text(response)


# Log errors
def error(update, context):
    print(f'Update {update} caused error {context.error}')


# Run the program
if __name__ == '__main__':
    updater = Updater(keys.BOT_TOKEN, use_context=True)
    dp = updater.dispatcher

    # Commands
    # dp.add_handler(CommandHandler('start', start_command))
    # dp.add_handler(CommandHandler('help', help_command))                                  
    # dp.add_handler(CommandHandler('custom', custom_command))                              
    dp.add_handler(CommandHandler('mybtc', my_btc))
    dp.add_handler(CommandHandler('myrvn', my_rvn))
    dp.add_handler(CommandHandler('stop', stop_notify))
    dp.add_handler(CommandHandler("startauto", start_auto_messaging))
    
    #chat id
    chat_id = 1987606009

    # Messages
    dp.add_handler(MessageHandler(Filters.text, handle_message))

    # Log all errors
    dp.add_error_handler(error)

    # Run the bot
    updater.start_polling(1.0)
    updater.idle()